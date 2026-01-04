"""
Export Channel Info to Excel Operator
Exports channel hydraulic and geometric information to Excel format (.xlsx)
and opens it with the default system application.
"""

import os
import platform
import subprocess
import sys
import tempfile
from datetime import datetime

import bpy
from bpy.props import BoolProperty, StringProperty
from bpy.types import Operator

# Global flag to track if we've already tried to install openpyxl this session
_openpyxl_install_attempted = False


def is_cadhy_channel(obj):
    """Check if object is a CADHY channel."""
    if not obj or obj.type != "MESH":
        return False
    ch = getattr(obj, "cadhy_channel", None)
    return ch is not None and ch.is_cadhy_object


def is_openpyxl_available() -> bool:
    """Check if openpyxl is installed."""
    try:
        import openpyxl  # noqa: F401

        return True
    except ImportError:
        return False


def install_openpyxl() -> tuple[bool, str]:
    """
    Install openpyxl package using pip.

    Returns:
        Tuple of (success, message)
    """
    global _openpyxl_install_attempted
    _openpyxl_install_attempted = True

    try:
        # Get the Python executable that Blender is using
        python_exe = sys.executable

        # Run pip install
        result = subprocess.run(
            [python_exe, "-m", "pip", "install", "openpyxl"],
            capture_output=True,
            text=True,
            timeout=60,  # 60 second timeout
        )

        if result.returncode == 0:
            # Verify installation
            try:
                import importlib

                importlib.invalidate_caches()
                import openpyxl  # noqa: F401

                return True, "openpyxl installed successfully!"
            except ImportError:
                return False, "Installation completed but import failed. Please restart Blender."
        else:
            error_msg = result.stderr or result.stdout or "Unknown error"
            return False, f"pip install failed: {error_msg[:200]}"

    except subprocess.TimeoutExpired:
        return False, "Installation timed out. Please try again or install manually."
    except FileNotFoundError:
        return False, "Python executable not found. Please install openpyxl manually."
    except Exception as e:
        return False, f"Installation error: {str(e)[:200]}"


def ensure_openpyxl() -> tuple[bool, str]:
    """
    Ensure openpyxl is available, installing it if necessary.

    Returns:
        Tuple of (available, message)
    """
    if is_openpyxl_available():
        return True, "openpyxl is available"

    # Try to install
    print("CADHY: openpyxl not found, attempting to install...")
    success, message = install_openpyxl()

    if success:
        print(f"CADHY: {message}")
        return True, message
    else:
        print(f"CADHY: {message}")
        return False, message


def open_file_with_default_app(filepath: str) -> bool:
    """
    Open a file with the system's default application.

    Args:
        filepath: Path to the file to open

    Returns:
        True if successful, False otherwise
    """
    try:
        system = platform.system()

        if system == "Darwin":  # macOS
            subprocess.Popen(["open", filepath])
        elif system == "Windows":
            os.startfile(filepath)
        else:  # Linux and others
            subprocess.Popen(["xdg-open", filepath])

        return True
    except Exception as e:
        print(f"Error opening file: {e}")
        return False


def export_channel_to_excel(obj, filepath: str, auto_install: bool = True) -> tuple[bool, str]:
    """
    Export channel data to Excel file.

    Args:
        obj: Blender object with cadhy_channel properties
        filepath: Output file path
        auto_install: If True, attempt to install openpyxl if not available

    Returns:
        Tuple of (success, actual_filepath_or_error_message)
    """
    try:
        # Check/install openpyxl
        if auto_install and not is_openpyxl_available():
            success, msg = ensure_openpyxl()
            if not success:
                # Fall back to CSV
                csv_result = export_channel_to_csv_fallback(obj, filepath)
                if csv_result:
                    csv_path = filepath[:-5] + ".csv" if filepath.lower().endswith(".xlsx") else filepath + ".csv"
                    return True, csv_path
                return False, "Failed to export (openpyxl install failed, CSV fallback failed)"

        # Try to import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Border, Font, PatternFill, Side
        except ImportError:
            # openpyxl not available, fall back to CSV
            csv_result = export_channel_to_csv_fallback(obj, filepath)
            if csv_result:
                csv_path = filepath[:-5] + ".csv" if filepath.lower().endswith(".xlsx") else filepath + ".csv"
                return True, csv_path
            return False, "openpyxl not available and CSV fallback failed"

        ch = obj.cadhy_channel

        # Create workbook
        wb = Workbook()

        # === SHEET 1: Summary ===
        ws_summary = wb.active
        ws_summary.title = "Channel Summary"

        # Styles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        section_font = Font(bold=True, size=11)
        section_fill = PatternFill(start_color="A0D2DB", end_color="A0D2DB", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws_summary["A1"] = "CADHY Channel Report"
        ws_summary["A1"].font = Font(bold=True, size=16)
        ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary["A3"] = f"Object: {obj.name}"

        row = 5

        # === GEOMETRY SECTION ===
        ws_summary[f"A{row}"] = "GEOMETRY"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = section_fill
        ws_summary[f"B{row}"].fill = section_fill
        ws_summary[f"C{row}"].fill = section_fill
        row += 1

        section_type_names = {
            "TRAP": "Trapezoidal",
            "RECT": "Rectangular",
            "TRI": "Triangular",
            "CIRC": "Semi-circular U",
            "PIPE": "Pipe",
        }

        geometry_data = [
            ("Section Type", section_type_names.get(ch.section_type, ch.section_type), ""),
            ("Bottom Width", f"{ch.bottom_width:.3f}", "m"),
            ("Side Slope", f"{ch.side_slope:.2f}", "H:V"),
            ("Height", f"{ch.height:.3f}", "m"),
            ("Freeboard", f"{ch.freeboard:.3f}", "m"),
            ("Total Height", f"{ch.height + ch.freeboard:.3f}", "m"),
            ("Lining Thickness", f"{ch.lining_thickness * 100:.1f}", "cm"),
        ]

        if ch.section_type == "TRAP":
            total_h = ch.height + ch.freeboard
            top_w = ch.bottom_width + 2 * ch.side_slope * total_h
            geometry_data.insert(3, ("Top Width", f"{top_w:.3f}", "m"))

        for param, value, unit in geometry_data:
            ws_summary[f"A{row}"] = param
            ws_summary[f"B{row}"] = value
            ws_summary[f"C{row}"] = unit
            ws_summary[f"A{row}"].border = thin_border
            ws_summary[f"B{row}"].border = thin_border
            ws_summary[f"C{row}"].border = thin_border
            row += 1

        row += 1

        # === PROFILE & SLOPE SECTION ===
        ws_summary[f"A{row}"] = "PROFILE & SLOPE"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = section_fill
        ws_summary[f"B{row}"].fill = section_fill
        ws_summary[f"C{row}"].fill = section_fill
        row += 1

        profile_data = [
            ("Total Length", f"{ch.total_length:.2f}", "m"),
            ("Slope", f"{ch.slope_percent:.4f}", "%"),
            ("Slope (m/m)", f"{ch.slope_avg:.6f}", "m/m"),
            ("Start Elevation", f"{ch.elevation_start:.2f}", "m"),
            ("End Elevation", f"{ch.elevation_end:.2f}", "m"),
            ("Elevation Drop", f"{ch.elevation_drop:.2f}", "m"),
        ]

        for param, value, unit in profile_data:
            ws_summary[f"A{row}"] = param
            ws_summary[f"B{row}"] = value
            ws_summary[f"C{row}"] = unit
            ws_summary[f"A{row}"].border = thin_border
            ws_summary[f"B{row}"].border = thin_border
            ws_summary[f"C{row}"].border = thin_border
            row += 1

        row += 1

        # === HYDRAULICS SECTION ===
        ws_summary[f"A{row}"] = "HYDRAULICS (Manning)"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = section_fill
        ws_summary[f"B{row}"].fill = section_fill
        ws_summary[f"C{row}"].fill = section_fill
        row += 1

        hydraulics_data = [
            ("Manning n", f"{ch.manning_n:.4f}", ""),
            ("Water Depth", f"{ch.height:.3f}", "m"),
            ("Hydraulic Area", f"{ch.hydraulic_area:.4f}", "m²"),
            ("Wetted Perimeter", f"{ch.wetted_perimeter:.4f}", "m"),
            ("Hydraulic Radius", f"{ch.hydraulic_radius:.4f}", "m"),
            ("Top Width (Water)", f"{ch.top_width_water:.3f}", "m"),
            ("Velocity", f"{ch.manning_velocity:.4f}", "m/s"),
            ("Discharge", f"{ch.manning_discharge:.4f}", "m³/s"),
            ("Discharge", f"{ch.manning_discharge * 1000:.2f}", "L/s"),
        ]

        for param, value, unit in hydraulics_data:
            ws_summary[f"A{row}"] = param
            ws_summary[f"B{row}"] = value
            ws_summary[f"C{row}"] = unit
            ws_summary[f"A{row}"].border = thin_border
            ws_summary[f"B{row}"].border = thin_border
            ws_summary[f"C{row}"].border = thin_border
            row += 1

        row += 1

        # === MESH STATISTICS ===
        ws_summary[f"A{row}"] = "MESH STATISTICS"
        ws_summary[f"A{row}"].font = section_font
        ws_summary[f"A{row}"].fill = section_fill
        ws_summary[f"B{row}"].fill = section_fill
        ws_summary[f"C{row}"].fill = section_fill
        row += 1

        mesh_data = [
            ("Vertices", f"{ch.mesh_vertices:,}", ""),
            ("Edges", f"{ch.mesh_edges:,}", ""),
            ("Faces", f"{ch.mesh_faces:,}", ""),
            ("Triangles", f"{ch.mesh_triangles:,}", ""),
            ("Volume", f"{ch.mesh_volume:.4f}", "m³"),
            ("Surface Area", f"{ch.mesh_surface_area:.4f}", "m²"),
            ("Manifold", "Yes" if ch.mesh_is_manifold else "No", ""),
            ("Watertight", "Yes" if ch.mesh_is_watertight else "No", ""),
            ("Non-Manifold Edges", f"{ch.mesh_non_manifold}", ""),
        ]

        for param, value, unit in mesh_data:
            ws_summary[f"A{row}"] = param
            ws_summary[f"B{row}"] = value
            ws_summary[f"C{row}"] = unit
            ws_summary[f"A{row}"].border = thin_border
            ws_summary[f"B{row}"].border = thin_border
            ws_summary[f"C{row}"].border = thin_border
            row += 1

        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 25
        ws_summary.column_dimensions["B"].width = 15
        ws_summary.column_dimensions["C"].width = 10

        # === SHEET 2: Data Table (for further analysis) ===
        ws_data = wb.create_sheet("Data Table")

        # Header row
        headers = ["Parameter", "Value", "Unit", "Category"]
        for col, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        # All data in table format
        all_data = []

        # Geometry
        for param, value, unit in geometry_data:
            all_data.append((param, value, unit, "Geometry"))

        # Profile
        for param, value, unit in profile_data:
            all_data.append((param, value, unit, "Profile"))

        # Hydraulics
        for param, value, unit in hydraulics_data:
            all_data.append((param, value, unit, "Hydraulics"))

        # Mesh
        for param, value, unit in mesh_data:
            all_data.append((param, value, unit, "Mesh"))

        for row_idx, (param, value, unit, category) in enumerate(all_data, 2):
            ws_data.cell(row=row_idx, column=1, value=param).border = thin_border

            # Try to convert to number for Excel
            try:
                num_value = float(value.replace(",", ""))
                ws_data.cell(row=row_idx, column=2, value=num_value).border = thin_border
            except (ValueError, AttributeError):
                ws_data.cell(row=row_idx, column=2, value=value).border = thin_border

            ws_data.cell(row=row_idx, column=3, value=unit).border = thin_border
            ws_data.cell(row=row_idx, column=4, value=category).border = thin_border

        # Adjust column widths
        ws_data.column_dimensions["A"].width = 25
        ws_data.column_dimensions["B"].width = 15
        ws_data.column_dimensions["C"].width = 10
        ws_data.column_dimensions["D"].width = 12

        # Save workbook
        if not filepath.lower().endswith(".xlsx"):
            filepath += ".xlsx"

        wb.save(filepath)
        return True, filepath

    except Exception as e:
        print(f"Excel export error: {e}")
        import traceback

        traceback.print_exc()
        return False, str(e)


def export_channel_to_csv_fallback(obj, filepath: str) -> bool:
    """
    Fallback to CSV if openpyxl is not available.

    Args:
        obj: Blender object with cadhy_channel properties
        filepath: Output file path (will be changed to .csv)

    Returns:
        True if successful
    """
    try:
        ch = obj.cadhy_channel

        # Change extension to CSV
        if filepath.lower().endswith(".xlsx"):
            filepath = filepath[:-5] + ".csv"
        elif not filepath.lower().endswith(".csv"):
            filepath += ".csv"

        lines = [
            "CADHY Channel Report",
            f"Generated,{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Object,{obj.name}",
            "",
            "Category,Parameter,Value,Unit",
            "",
            "Geometry,Section Type," + ch.section_type + ",",
            f"Geometry,Bottom Width,{ch.bottom_width:.3f},m",
            f"Geometry,Side Slope,{ch.side_slope:.2f},H:V",
            f"Geometry,Height,{ch.height:.3f},m",
            f"Geometry,Freeboard,{ch.freeboard:.3f},m",
            f"Geometry,Total Height,{ch.height + ch.freeboard:.3f},m",
            "",
            f"Profile,Total Length,{ch.total_length:.2f},m",
            f"Profile,Slope,{ch.slope_percent:.4f},%",
            f"Profile,Start Elevation,{ch.elevation_start:.2f},m",
            f"Profile,End Elevation,{ch.elevation_end:.2f},m",
            f"Profile,Elevation Drop,{ch.elevation_drop:.2f},m",
            "",
            f"Hydraulics,Manning n,{ch.manning_n:.4f},",
            f"Hydraulics,Hydraulic Area,{ch.hydraulic_area:.4f},m²",
            f"Hydraulics,Wetted Perimeter,{ch.wetted_perimeter:.4f},m",
            f"Hydraulics,Hydraulic Radius,{ch.hydraulic_radius:.4f},m",
            f"Hydraulics,Velocity,{ch.manning_velocity:.4f},m/s",
            f"Hydraulics,Discharge,{ch.manning_discharge:.4f},m³/s",
            "",
            f"Mesh,Vertices,{ch.mesh_vertices},",
            f"Mesh,Faces,{ch.mesh_faces},",
            f"Mesh,Volume,{ch.mesh_volume:.4f},m³",
            f"Mesh,Manifold,{'Yes' if ch.mesh_is_manifold else 'No'},",
            f"Mesh,Watertight,{'Yes' if ch.mesh_is_watertight else 'No'},",
        ]

        with open(filepath, "w", encoding="utf-8") as f:
            f.write("\n".join(lines))

        return True

    except Exception as e:
        print(f"CSV fallback export error: {e}")
        return False


class CADHY_OT_ExportChannelExcel(Operator):
    """Export channel information to Excel and open with default application"""

    bl_idname = "cadhy.export_channel_excel"
    bl_label = "Export to Excel"
    bl_description = "Export channel data to Excel (.xlsx) and open with default spreadsheet application"
    bl_options = {"REGISTER"}

    filepath: StringProperty(
        name="File Path",
        description="Path to export file",
        subtype="FILE_PATH",
    )

    open_after_export: BoolProperty(
        name="Open After Export",
        description="Open the file with default application after export",
        default=True,
    )

    @classmethod
    def poll(cls, context):
        return is_cadhy_channel(context.active_object)

    def execute(self, context):
        obj = context.active_object

        # Ensure .xlsx extension
        filepath = self.filepath
        if not filepath.lower().endswith(".xlsx"):
            filepath += ".xlsx"

        # Check if openpyxl is available, if not try to install
        if not is_openpyxl_available():
            self.report({"INFO"}, "Installing Excel support (openpyxl)... Please wait.")

        # Export to Excel (will auto-install openpyxl if needed)
        success, result = export_channel_to_excel(obj, filepath, auto_install=True)

        if success:
            actual_path = result

            # Check if it's a CSV (fallback)
            if actual_path.endswith(".csv"):
                self.report({"WARNING"}, "Exported as CSV (Excel support not available)")
            else:
                self.report({"INFO"}, f"Exported to Excel: {os.path.basename(actual_path)}")

            # Open with default application
            if self.open_after_export:
                if open_file_with_default_app(actual_path):
                    self.report({"INFO"}, "Opening file...")
                else:
                    self.report({"WARNING"}, "Could not open file automatically")

            return {"FINISHED"}
        else:
            self.report({"ERROR"}, f"Export failed: {result}")
            return {"CANCELLED"}

    def invoke(self, context, event):
        """Show file browser."""
        obj = context.active_object

        if not self.filepath:
            # Try to get export path from settings
            settings = context.scene.cadhy
            export_dir = bpy.path.abspath(settings.export_path)

            if not os.path.exists(export_dir):
                export_dir = bpy.path.abspath("//")

            if not os.path.exists(export_dir):
                export_dir = tempfile.gettempdir()

            # Generate filename from object name
            safe_name = "".join(c for c in obj.name if c.isalnum() or c in "._- ")
            self.filepath = os.path.join(export_dir, f"{safe_name}_report.xlsx")

        context.window_manager.fileselect_add(self)
        return {"RUNNING_MODAL"}

    def draw(self, context):
        """Draw export options."""
        layout = self.layout
        layout.prop(self, "open_after_export")
